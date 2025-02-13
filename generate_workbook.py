import openpyxl

def clash_reprt():
    # ...existing code if any...
    return [
        ['Event', 'Time', 'Location'],
        ['Exam Clash 1', '09:00', 'Room A'],
        ['Exam Clash 2', '11:00', 'Room B']
    ]

def make_exam_slot():
    # ...existing code if any...
    return [
        ['Slot', 'Exam Name', 'Duration'],
        ['Morning Slot', 'Math', '2 hrs'],
        ['Afternoon Slot', 'English', '1.5 hrs']
    ]

def generate_workbook():
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create and populate Clash Report sheet
    clash_sheet = wb.create_sheet(title='Clash Report')
    for row in clash_reprt():
        clash_sheet.append(row)
    
    # Create and populate Exam Slot sheet
    exam_sheet = wb.create_sheet(title='Exam Slot')
    for row in make_exam_slot():
        exam_sheet.append(row)
    
    wb.save('report_workbook.xlsx')

if __name__ == '__main__':
    generate_workbook()
